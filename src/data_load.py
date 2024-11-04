import re
import pandas as pd
import configparser
from openpyxl import load_workbook

########################################################
###########Load Conceptual Mapping Sheet################
########################################################

class dataLoader():
    def __init__(self, file_path, config_path, cm_version):
        # load the configuration file
        
        self.cm_version = cm_version
        self.config = configparser.ConfigParser()
        self.config.read(config_path)
        self.file_path = file_path

        # load the data
        self.sheet_names = {}
        for i in ["metadata_sheet_name", "rules_sheet_name", "cl1_sheet_name"]:
            self.sheet_names[i] = self.config.get(self.cm_version, i)

        self.workbook = load_workbook(self.file_path, data_only=True)


    def load(self):

        if self.sheet_names['metadata_sheet_name'] != "":
            metaData_info = self.load_metadata(self.sheet_names['metadata_sheet_name'])
        else: metaData_info = ""

        if self.sheet_names['rules_sheet_name'] != "":
            class_paths, property_paths, field_XPaths, field_id = self.load_rules(self.sheet_names['rules_sheet_name'])
        else: raise ValueError("Rules Information is not provided in the configuration file")
        
        if self.sheet_names['cl1_sheet_name'] != "":
            cl1 = self.load_controlled_list(self.sheet_names['cl1_sheet_name'])
        else: cl1 = {"CL1":{}}

        # release the memory
        del self.workbook

        return metaData_info, class_paths, property_paths, field_XPaths, cl1, field_id

    def _split_cell(self, cell):
        match = re.match(r"([A-Z]+)([0-9]+)", cell, re.I)
        if match:
            items = match.groups()
            return items[0], int(items[1])
        return None, None

    def _get_cell_value(self, sheet_name, cell):
        sheet = self.workbook[sheet_name]
        return sheet[cell].value

    def _get_column_values(self, sheet_name, start_cell):
        sheet = self.workbook[sheet_name]
       
        col, start_row = self._split_cell(start_cell)
        
        values = []
        for row in range(start_row, sheet.max_row + 1):
            cell_value = sheet[f"{col}{row}"].value
            values.append(cell_value)
        
        return values

    def _get_rules_values(self, sheet_name, start_cell1, start_cell2, start_cell3, start_cell4):

        sheet = self.workbook[sheet_name]

        col1, start_row1 = self._split_cell(start_cell1)
        col2, start_row2 = self._split_cell(start_cell2)
        col3, start_row3 = self._split_cell(start_cell3)
        col4, start_row4 = self._split_cell(start_cell4)

        if col1 is None or start_row1 is None or col2 is None or start_row2 is None or col3 is None or start_row3 is None or col4 is None or start_row4 is None:
            print("Error: Invalid cell value")
            return [], [], [], []

        l1 = []
        l2 = []
        l3 = []
        l4 = []
        max_row = max(sheet.max_row, start_row1, start_row2, start_row3, start_row4)
        for row in range(max(start_row1, start_row2, start_row3, start_row4), max_row + 1):
            cell_value1 = sheet[f"{col1}{row}"].value
            cell_value2 = sheet[f"{col2}{row}"].value
            cell_value3 = sheet[f"{col3}{row}"].value
            cell_value4 = sheet[f"{col4}{row}"].value
            if cell_value1 is not None and cell_value2 is not None: # Xpath and ID are optional
                l1.append(cell_value1)
                l2.append(cell_value2)
                l3.append(cell_value3)
                l4.append(cell_value4)

        return l1, l2, l3, l4
        
    def load_metadata(self, sheet_name):

        identifier_cell = self.config[self.cm_version].get('Identifier', '')
        description_cell = self.config[self.cm_version].get('Description', '')
        mapping_version_cell = self.config[self.cm_version].get('Mapping_Version', '')
        epo_version_cell = self.config[self.cm_version].get('EPO_version', '')
        base_xpath_cell = self.config[self.cm_version].get('baseXpath', '')

        identifier = self._get_cell_value(sheet_name, identifier_cell) if identifier_cell != '' else ''
        description = self._get_cell_value(sheet_name, description_cell) if description_cell != '' else ''
        mapping_version = self._get_cell_value(sheet_name, mapping_version_cell) if mapping_version_cell != '' else ''
        epo_version = self._get_cell_value(sheet_name, epo_version_cell) if epo_version_cell != '' else ''
        base_xpath = self._get_cell_value(sheet_name, base_xpath_cell) if base_xpath_cell != '' else ''

        metaData_info = f"""# The SHACL shapes graph is automatic translated from the Conceptual Mapping below: 
        # Identifier: {identifier}, 
        # Description: {description},
        # Mapping Version: {mapping_version}
        # EPO version: {epo_version}."""

        return metaData_info

    def load_rules(self, sheet_name):

        class_path_cell = self.config[self.cm_version].get('Class_path', '')
        property_path_cell = self.config[self.cm_version].get('Property_path', '')
        field_XPath_cell = self.config[self.cm_version].get('Field_XPath', '')
        field_id_cell = self.config[self.cm_version].get('Field_ID', '')

        class_paths, property_paths, field_XPaths, field_id = self._get_rules_values(sheet_name, class_path_cell, property_path_cell, field_XPath_cell, field_id_cell)

        return class_paths, property_paths, field_XPaths, field_id

    def load_controlled_list(self, sheet_name):
        xml_path_fragment_cell = self.config[self.cm_version].get('XML_PATH_Fragment', '')
        mapping_reference_cell = self.config[self.cm_version].get('Mapping_Reference', '')

        xml_path_fragment = self._get_column_values(sheet_name, xml_path_fragment_cell) if xml_path_fragment_cell != '' else []
        mapping_reference = self._get_column_values(sheet_name, mapping_reference_cell) if mapping_reference_cell != '' else []

        cl1Df_dict = dict(zip(xml_path_fragment, mapping_reference))
        return {'CL1':cl1Df_dict}